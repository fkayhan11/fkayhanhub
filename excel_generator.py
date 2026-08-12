import json
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_excel(json_path, output_path):
    # Load input data
    with open(json_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    title = config.get("title", "YİYECEK & RESTORAN KONSEPT DETAYLI RAPORU")
    subtitle = config.get("subtitle", "Tüm veri kümesini ve geri bildirim analizlerini kapsayan tam liste.")
    sheet_name = config.get("sheet_name", "Yiyecek Konseptleri Raporu")
    headers = config.get("headers", [
        "Konsept (Concept)",
        "Hacim (Volume)",
        "Trend",
        "Pozitif Hacim (Pos. Vol)",
        "Pozitif % (Pos. %)",
        "Negatif Hacim (Neg. Vol)",
        "Negatif % (Neg. %)"
    ])
    data = config.get("data", [])

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Ensure gridlines are visible
    ws.views.sheetView[0].showGridLines = True

    # Title in B4 (Calibri 16pt Bold, Deep Blue #1F497D)
    ws['B4'] = title
    ws['B4'].font = Font(name="Calibri", size=16, bold=True, color="1F497D")

    # Subtitle in B5 (Calibri 11pt Italic, Dark Gray #595959)
    ws['B5'] = subtitle
    ws['B5'].font = Font(name="Calibri", size=11, italic=True, color="595959")

    # Write headers to B7:H7
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

    # Styling helpers
    thin_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Write data starting at row 8
    for row_idx, item in enumerate(data, start=8):
        # Expected item: [concept, volume, trend, pos_vol, neg_vol]
        concept, volume, trend, pos_vol, neg_vol = item
        
        # Column B: Concept (Left)
        c_cell = ws.cell(row=row_idx, column=2, value=concept)
        c_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Column C: Volume (Right)
        v_cell = ws.cell(row=row_idx, column=3, value=volume)
        v_cell.number_format = '#,##0'
        v_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Column D: Trend (Right, custom sign format)
        t_cell = ws.cell(row=row_idx, column=4, value=trend)
        t_cell.number_format = '+0;-0;0'
        t_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Column E: Pos Vol (Right)
        pv_cell = ws.cell(row=row_idx, column=5, value=pos_vol)
        pv_cell.number_format = '#,##0'
        pv_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Column F: Pos % (Formula)
        pp_cell = ws.cell(row=row_idx, column=6, value=f"=IFERROR(E{row_idx}/C{row_idx}, 0)")
        pp_cell.number_format = '0.0%'
        pp_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Column G: Neg Vol (Right)
        nv_cell = ws.cell(row=row_idx, column=7, value=neg_vol)
        nv_cell.number_format = '#,##0'
        nv_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Column H: Neg % (Formula)
        np_cell = ws.cell(row=row_idx, column=8, value=f"=IFERROR(G{row_idx}/C{row_idx}, 0)")
        np_cell.number_format = '0.0%'
        np_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Alternating row background (zebra fill)
        current_fill = white_fill if (row_idx % 2 == 0) else zebra_fill
        
        for col_idx in range(2, 9):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Calibri", size=11)
            cell.fill = current_fill
            cell.border = thin_border

    # Set row heights
    ws.row_dimensions[4].height = 28  # Title
    ws.row_dimensions[5].height = 18  # Subtitle
    ws.row_dimensions[6].height = 15  # Spacing
    ws.row_dimensions[7].height = 28  # Header
    for r in range(8, 8 + len(data)):
        ws.row_dimensions[r].height = 20

    # Set column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 24
    ws.column_dimensions['H'].width = 18

    # Save
    wb.save(output_path)
    print(f"Successfully generated: {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 excel_generator.py <data_json_path> <output_xlsx_path>")
        sys.exit(1)
    generate_excel(sys.argv[1], sys.argv[2])
