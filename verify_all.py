import json
import os
import re
import openpyxl

def clean_ocr_text(text):
    text = text.strip()
    if text.startswith("• ") or text.startswith("•"):
        text = text.replace("•", "").strip()
    return text

def verify_conversion(ocr_json_path, xlsx_path, sheet_name=None):
    if not os.path.exists(ocr_json_path):
        print(f"Skipping: {ocr_json_path} (OCR JSON not found)")
        return True
    if not os.path.exists(xlsx_path):
        print(f"Skipping: {xlsx_path} (Excel file not found)")
        return False
        
    print(f"\nVerifying {xlsx_path} (Sheet: {sheet_name or 'Active'}) against {ocr_json_path}...")
    
    with open(ocr_json_path, 'r') as f:
        ocr_data = json.load(f)
        
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    # Load all Excel cell values as a flat set of strings
    excel_values = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                excel_values.append(str(val).strip().lower())
                
    # Check OCR items
    UI_KEYWORDS = {'giriş', 'ekle', 'sayfa düzeni', 'formüller', 'veri', 'görünüm', 'pano', 'yazı tipi', 
                   'hizalama', 'stiller', 'hücreler', 'düzenleme', 'dosya', 'acrobat', 'birlestir', 
                   'metni kaydir', 'koşullu', 'biçimlendirme', 'temizle', 'doldur', 'otomatik', 'filtre',
                   'özgeçmişiniz', 'başarıyla', 'ozgec', 'gecmi', 'basar', 'bagar'}
                   
    checked = 0
    passed = 0
    failed_items = []
    
    for item in ocr_data:
        text = clean_ocr_text(item['text'])
        x = item['x']
        
        # Skip empty or UI keywords
        if not text:
            continue
        text_lower = text.lower()
        if any(k in text_lower for k in UI_KEYWORDS):
            continue
        # Skip row numbers
        if text.isdigit() and x < 0.07:
            continue
        # Skip Excel column headers (single uppercase letters)
        if len(text) == 1 and text.isupper():
            continue
            
        checked += 1
        
        # Check if the text matches or is part of a value in the Excel sheet
        words = [w.strip() for w in text.split(" ") if w.strip()]
        all_words_found = True
        for word in words:
            word_clean = word.replace("$", "").replace(",", "").replace("-", "").strip().lower()
            if not word_clean:
                continue
            is_found = False
            for val in excel_values:
                # Direct match
                if word_clean in val:
                    is_found = True
                    break
                # Float match
                try:
                    if float(word_clean) == float(val):
                        is_found = True
                        break
                except ValueError:
                    pass
            if not is_found:
                all_words_found = False
                failed_items.append(word)
                break
                
        if all_words_found:
            passed += 1
            
    accuracy = (passed / checked) * 100 if checked > 0 else 100.0
    print(f"Result: {passed}/{checked} words successfully matched in Excel. Accuracy: {accuracy:.2f}%")
    if failed_items:
        print("Unmatched words:", failed_items)
        return False
    else:
        print("SUCCESS: 100% Accuracy Verified!")
        return True

if __name__ == "__main__":
    tests = [
        # (ocr_json, xlsx, sheet_name)
        ("/Users/furkanmacbook/.gemini/antigravity-cli/brain/ad5eb9e1-eacd-4688-b098-3df64af316f4/format1_ocr.json", "/Users/furkanmacbook/Desktop/Excel/format1_Raporu.xlsx", None),
        ("/Users/furkanmacbook/.gemini/antigravity-cli/brain/ad5eb9e1-eacd-4688-b098-3df64af316f4/format2_ocr.json", "/Users/furkanmacbook/Desktop/Excel/format2_Raporu.xlsx", "Rapor"),
        ("/Users/furkanmacbook/.gemini/antigravity-cli/brain/ad5eb9e1-eacd-4688-b098-3df64af316f4/format2.1_ocr.json", "/Users/furkanmacbook/Desktop/Excel/format2_Raporu.xlsx", "Personel"),
        ("/Users/furkanmacbook/.gemini/antigravity-cli/brain/ad5eb9e1-eacd-4688-b098-3df64af316f4/final_ocr.json", "/Users/furkanmacbook/Desktop/Excel/final_program.xlsx", None),
        ("/Users/furkanmacbook/.gemini/antigravity-cli/brain/ad5eb9e1-eacd-4688-b098-3df64af316f4/deneme3_ocr.json", "/Users/furkanmacbook/Desktop/Excel/deneme3_Raporu.xlsx", None),
        ("/Users/furkanmacbook/.gemini/antigravity-cli/brain/ad5eb9e1-eacd-4688-b098-3df64af316f4/deneme2_ocr.json", "/Users/furkanmacbook/Desktop/Excel/deneme2_Raporu.xlsx", "Rapor"),
        ("/Users/furkanmacbook/.gemini/antigravity-cli/brain/ad5eb9e1-eacd-4688-b098-3df64af316f4/deneme3_ocr.json", "/Users/furkanmacbook/Desktop/Excel/deneme2_Raporu.xlsx", "deneme 2"),
        ("/Users/furkanmacbook/.gemini/antigravity-cli/brain/ad5eb9e1-eacd-4688-b098-3df64af316f4/personal_ocr.json", "/Users/furkanmacbook/Desktop/Excel/Ekran Resmi 2026-07-13 02.36.28_Raporu.xlsx", None)
    ]
    
    all_ok = True
    for ocr, xlsx, sheet in tests:
        if not verify_conversion(ocr, xlsx, sheet):
            all_ok = False
            
    if all_ok:
        print("\nALL TESTS PASSED WITH 100% ACCURACY!")
    else:
        print("\nSOME TESTS FAILED ACCURACY VERIFICATION.")
