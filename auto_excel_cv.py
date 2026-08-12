import sys
import os
import json
import re
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import cv2
    import numpy as np
    from sklearn.cluster import KMeans
except ImportError:
    print("Gerekli kütüphaneler (opencv-python, numpy, scikit-learn) yüklenmemiş.")
    sys.exit(1)

def run_swift_ocr(image_path):
    """Run the local Swift OCR to get precise text and bounding boxes."""
    swift_script = os.path.join(os.path.dirname(__file__), "ocr.swift")
    try:
        result = subprocess.run(['swift', swift_script, image_path], 
                                capture_output=True, text=True, check=True)
        raw_json = result.stdout.strip()
        data = json.loads(raw_json)
        return data
    except Exception as e:
        print(f"Error running OCR: {e}")
        return []

def rgb_to_hex(rgb):
    """Convert RGB tuple to HEX string for OpenPyXL"""
    return f"{int(rgb[0]):02X}{int(rgb[1]):02X}{int(rgb[2]):02X}"

def extract_colors(roi):
    """Extract background and text colors from an image region using K-Means clustering."""
    if roi.size == 0:
        return "FFFFFF", "000000", False

    # Reshape the image to be a list of pixels
    pixels = roi.reshape((-1, 3))
    
    # We expect 2 main colors: Background and Text
    try:
        kmeans = KMeans(n_clusters=2, n_init=3, max_iter=100)
        kmeans.fit(pixels)
        
        # Count pixels in each cluster
        counts = np.bincount(kmeans.labels_)
        
        # The background is usually the cluster with MORE pixels
        bg_cluster = np.argmax(counts)
        text_cluster = 1 - bg_cluster
        
        bg_color = kmeans.cluster_centers_[bg_cluster]
        text_color = kmeans.cluster_centers_[text_cluster]
        
        # Calculate if text is bold by checking the ratio of text pixels
        # If text takes up >25% of the box, it might be bold
        text_ratio = counts[text_cluster] / float(len(pixels))
        is_bold = text_ratio > 0.25
        
        # Convert BGR (OpenCV default) to RGB, then to HEX
        bg_rgb = bg_color[::-1]
        bg_hex = rgb_to_hex(bg_rgb)
        
        # Calculate luminance of background to ensure text is readable
        luminance = 0.299 * bg_rgb[0] + 0.587 * bg_rgb[1] + 0.114 * bg_rgb[2]
        
        # Force text color to Black or White for maximum readability
        if luminance > 128:
            text_hex = "000000"  # Dark text on light background
        else:
            text_hex = "FFFFFF"  # Light text on dark background
        
        return bg_hex, text_hex, is_bold
    except Exception:
        return "FFFFFF", "000000", False

def process_image_advanced(image_path, target_xlsx):
    print(f"[*] Gelişmiş Görüntü İşleme (OpenCV) başlatılıyor: {image_path}")
    
    # 1. Mac Vision OCR'ı çalıştırıp kutuları al
    ocr_data = run_swift_ocr(image_path)
    if not ocr_data:
        raise Exception("Görselden hiçbir metin okunamadı veya görsel boş.")
        
    # 2. Resmi OpenCV ile Oku
    img = cv2.imread(image_path)
    if img is None:
        raise Exception("Görsel OpenCV tarafından açılamadı, dosya bozuk olabilir.")
        
    img_h, img_w, _ = img.shape
    
    print("[*] Piksel bazlı renk ve font analizi yapılıyor...")
    # 3. Her hücre için renk analizi yap
    for item in ocr_data:
        x_norm, y_norm = item['x'], item['y']
        w_norm, h_norm = item['width'], item['height']
        
        # Normalize koordinatları piksele çevir (OCR kutuları bazen dar olabilir, biraz esnetelim)
        x1 = max(0, int(x_norm * img_w) - 2)
        y1 = max(0, int(y_norm * img_h) - 2)
        x2 = min(img_w, int((x_norm + w_norm) * img_w) + 2)
        y2 = min(img_h, int((y_norm + h_norm) * img_h) + 2)
        
        roi = img[y1:y2, x1:x2]
        
        bg_hex, text_hex, is_bold = extract_colors(roi)
        
        item['bg_color'] = bg_hex
        item['text_color'] = text_hex
        item['is_bold'] = is_bold
        item['pixel_y'] = y1 # Gerçek piksel Y
        item['pixel_x'] = x1 # Gerçek piksel X
        item['pixel_h'] = y2 - y1

    # 4. Satır ve Sütunları Yapay Zeka/Matematiksel Gruplama (Eski sisteme benzer ama piksel tabanlı)
    # Sort by Y, then X
    ocr_data.sort(key=lambda x: (x['pixel_y'], x['pixel_x']))
    
    rows = []
    current_row = []
    current_y = -1
    
    # Average height to determine row boundary tolerance
    avg_h = sum(item['pixel_h'] for item in ocr_data) / len(ocr_data)
    y_tolerance = avg_h * 0.5
    
    for item in ocr_data:
        if current_y == -1:
            current_y = item['pixel_y']
            current_row.append(item)
        elif abs(item['pixel_y'] - current_y) < y_tolerance:
            current_row.append(item)
        else:
            current_row.sort(key=lambda x: x['pixel_x'])
            rows.append(current_row)
            current_row = [item]
            current_y = item['pixel_y']
            
    if current_row:
        current_row.sort(key=lambda x: x['pixel_x'])
        rows.append(current_row)
        
    print(f"[*] Toplam {len(rows)} satır tespit edildi. Excel oluşturuluyor...")
    
    # 5. Excel'e Mükemmel Çıktıyı Bas
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analiz"
    
    # Sütunları X eksenine göre hizalamak için dinamik eşleştirme
    # Bütün satırlardaki x koordinatlarını toplayıp sütunları belirleyelim
    all_x = [item['pixel_x'] for row in rows for item in row]
    # Basit bir kümeleme yapalım
    cols = []
    for x in sorted(all_x):
        if not cols:
            cols.append(x)
        elif x - cols[-1] > avg_h * 2: # Tolerance for column separation
            cols.append(x)
            
    for r_idx, row in enumerate(rows, start=1):
        for item in row:
            # En yakın sütunu bul
            col_idx = 1
            min_diff = float('inf')
            for i, cx in enumerate(cols, start=1):
                diff = abs(item['pixel_x'] - cx)
                if diff < min_diff:
                    min_diff = diff
                    col_idx = i
                    
            cell = ws.cell(row=r_idx, column=col_idx, value=item['text'])
            
            # STYLING (Mükemmellik Aşaması)
            # Arka plan rengi
            cell.fill = PatternFill(start_color=item['bg_color'], end_color=item['bg_color'], fill_type="solid")
            # Yazı rengi ve kalınlık
            cell.font = Font(name="Arial", size=11, bold=item['is_bold'], color=item['text_color'])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Sütun genişliğini ayarla
            col_letter = get_column_letter(col_idx)
            current_width = ws.column_dimensions[col_letter].width
            if current_width is None or len(item['text']) > current_width:
                ws.column_dimensions[col_letter].width = len(item['text']) * 1.2
                
    # Kenarlıklar
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for r in ws.iter_rows(min_row=1, max_row=len(rows), min_col=1, max_col=len(cols)):
        for c in r:
            c.border = border
            
    wb.save(target_xlsx)
    print(f"[+] Mükemmel kopya oluşturuldu: {target_xlsx}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Kullanım: python3 auto_excel_cv.py <resim_yolu> <hedef_excel_yolu>")
        sys.exit(1)
    process_image_advanced(sys.argv[1], sys.argv[2])
