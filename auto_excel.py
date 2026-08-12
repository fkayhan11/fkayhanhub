import sys
import os
import json
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_int(text):
    # Clean text to extract integer
    clean = text.replace(" ", "").replace("+", "").replace(",", "").replace(".", "")
    # Check if there is a minus sign
    is_neg = "-" in text
    clean = "".join(c for c in clean if c.isdigit())
    if not clean:
        return 0
    val = int(clean)
    return -val if is_neg else val

def process_image(image_path):
    print(f"Processing image: {image_path}")
    base_dir = os.path.dirname(os.path.abspath(image_path))
    base_name = os.path.splitext(os.path.basename(image_path))[0]
    
    # 1. Convert to PNG if HEIC/HEIF
    ext = os.path.splitext(image_path)[1].lower()
    temp_png = os.path.join(base_dir, f"temp_{base_name}.png")
    
    if ext in ['.heic', '.heif']:
        print("HEIC format detected. Converting to PNG...")
        cmd = ["sips", "-s", "format", "png", image_path, "--out", temp_png]
        subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        ocr_target = temp_png
    else:
        ocr_target = image_path
        
    # 2. Run Native OCR binary
    ocr_binary = os.path.join(base_dir, "ocr")
    if not os.path.exists(ocr_binary):
        ocr_binary = "/Users/furkanmacbook/Desktop/Excel/ocr"
        
    print("Running OCR text recognition...")
    cmd = [ocr_binary, ocr_target]
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    # Clean up temp png if created
    if os.path.exists(temp_png):
        os.remove(temp_png)
        
    if result.returncode != 0:
        print("Error running OCR script:", result.stderr)
        return
        
    try:
        ocr_data = json.loads(result.stdout)
    except json.JSONDecodeError:
        print("Error parsing OCR JSON output. Output was:", result.stdout)
        return

    # 3. Parse table structure from coordinates
    # Apple Vision coordinates: y is 0 at bottom, 1 at top. x is 0 at left, 1 at right.
    
    # We want to find:
    # - Title: y > 0.68, x ≈ 0.04
    # - Subtitle: y ≈ 0.66 to 0.70
    # - Header row (for columns)
    # - Data rows: y < 0.65
    
    # Step A: Identify title and subtitle
    title = "YİYECEK & RESTORAN KONSEPT DETAYLI RAPORU"
    subtitle = "Tüm veri kümesini ve geri bildirim analizlerini kapsayan tam liste."
    
    title_candidates = [i for i in ocr_data if i['y'] > 0.68 and i['x'] < 0.20 and len(i['text']) > 15]
    if title_candidates:
        # Sort by y descending (highest in image)
        title_candidates.sort(key=lambda i: -i['y'])
        title = title_candidates[0]['text']
        if len(title_candidates) > 1:
            subtitle = title_candidates[1]['text']
            
    print(f"Extracted Title: {title}")
    print(f"Extracted Subtitle: {subtitle}")

    # Step B: Calculate camera tilt slope using grid search
    cells = [i for i in ocr_data if 0.03 <= i['x'] <= 0.95 and 0.15 <= i['y'] <= 0.68]
    best_slope = -0.03
    max_alignments = 0
    
    # Try slopes from -0.06 to 0.01 with step 0.001
    for slope in [s / 1000.0 for s in range(-60, 10)]:
        y_projs = [c['y'] - slope * c['x'] for c in cells]
        alignments = 0
        for i in range(len(cells)):
            for j in range(i + 1, len(cells)):
                if abs(cells[i]['x'] - cells[j]['x']) > 0.15:
                    if abs(y_projs[i] - y_projs[j]) < 0.006:
                        alignments += 1
        if alignments > max_alignments:
            max_alignments = alignments
            best_slope = slope
            
    print(f"Calculated camera tilt slope: {best_slope:.4f}")
    
    # Define projection function to align cells horizontally
    def get_y_proj(item):
        return item['y'] - best_slope * item['x']

    # Step C: Group items into rows using projected Y-coordinates
    col_ranges = {
        "concept": (0.02, 0.30),
        "volume": (0.36, 0.44),
        "trend": (0.45, 0.52),
        "pos_vol": (0.53, 0.62),
        "neg_vol": (0.75, 0.84)
    }

    # Find all concepts (leftmost column text blocks) below headers
    concepts = [i for i in ocr_data if col_ranges["concept"][0] <= i['x'] <= col_ranges["concept"][1] and get_y_proj(i) < 0.66]
    
    # Sort concepts by projected Y coordinate descending (top of sheet to bottom)
    concepts.sort(key=lambda i: -get_y_proj(i))
    
    table_data = []
    print(f"Found {len(concepts)} rows. Extracting column values using tilt correction...")
    
    for concept_item in concepts:
        c_proj_y = get_y_proj(concept_item)
        c_text = concept_item['text']
        
        # Skip headers or sheet tabs
        if "konsept" in c_text.lower() or "concept" in c_text.lower() or "konseptleri" in c_text.lower():
            continue
            
        # Find cells in this row (within a tight projected Y tolerance of 0.008)
        row_cells = [i for i in ocr_data if abs(get_y_proj(i) - c_proj_y) <= 0.008]
        
        volume = 0
        trend = 0
        pos_vol = 0
        neg_vol = 0
        
        for cell in row_cells:
            cx = cell['x']
            text = cell['text']
            
            if col_ranges["volume"][0] <= cx <= col_ranges["volume"][1]:
                volume = parse_int(text)
            elif col_ranges["trend"][0] <= cx <= col_ranges["trend"][1]:
                trend = parse_int(text)
            elif col_ranges["pos_vol"][0] <= cx <= col_ranges["pos_vol"][1]:
                pos_vol = parse_int(text)
            elif col_ranges["neg_vol"][0] <= cx <= col_ranges["neg_vol"][1]:
                neg_vol = parse_int(text)
                
        table_data.append((c_text, volume, trend, pos_vol, neg_vol))

    # 4. Generate the styled Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yiyecek Konseptleri Raporu"
    ws.views.sheetView[0].showGridLines = True

    # Title Style
    ws['B4'] = title
    ws['B4'].font = Font(name="Calibri", size=16, bold=True, color="1F497D")

    # Subtitle Style
    ws['B5'] = subtitle
    ws['B5'].font = Font(name="Calibri", size=11, italic=True, color="595959")

    # Headers
    headers = [
        "Konsept (Concept)",
        "Hacim (Volume)",
        "Trend",
        "Pozitif Hacim (Pos. Vol)",
        "Pozitif % (Pos. %)",
        "Negatif Hacim (Neg. Vol)",
        "Negatif % (Neg. %)"
    ]

    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

    # Styling helpers
    thin_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row_idx, item in enumerate(table_data, start=8):
        concept, volume, trend, pos_vol, neg_vol = item
        
        # Concept (Column B)
        c_cell = ws.cell(row=row_idx, column=2, value=concept)
        c_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Volume (Column C)
        v_cell = ws.cell(row=row_idx, column=3, value=volume)
        v_cell.number_format = '#,##0'
        v_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Trend (Column D - Centered)
        t_cell = ws.cell(row=row_idx, column=4, value=trend)
        t_cell.number_format = '+0;-0;0'
        t_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Pos Vol (Column E)
        pv_cell = ws.cell(row=row_idx, column=5, value=pos_vol)
        pv_cell.number_format = '#,##0'
        pv_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Pos % (Column F - Formula)
        pp_cell = ws.cell(row=row_idx, column=6, value=f"=IFERROR(E{row_idx}/C{row_idx}, 0)")
        pp_cell.number_format = '0.0%'
        pp_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Neg Vol (Column G)
        nv_cell = ws.cell(row=row_idx, column=7, value=neg_vol)
        nv_cell.number_format = '#,##0'
        nv_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Neg % (Column H - Formula)
        np_cell = ws.cell(row=row_idx, column=8, value=f"=IFERROR(G{row_idx}/C{row_idx}, 0)")
        np_cell.number_format = '0.0%'
        np_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Alternating row background (zebra fill)
        current_fill = white_fill if (row_idx % 2 == 0) else zebra_fill
        
        for col_idx in range(2, 9):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Calibri", size=11)
            cell.fill = current_fill
            cell.border = thin_border

    # Set row heights
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 15
    ws.row_dimensions[7].height = 28
    for r in range(8, 8 + len(table_data)):
        ws.row_dimensions[r].height = 20

    # Set column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 24
    ws.column_dimensions['H'].width = 18

    # Save output Excel file
    output_xlsx = os.path.join(base_dir, f"{base_name}_Raporu.xlsx")
    wb.save(output_xlsx)
    print(f"\nSUCCESS! Created styled Excel file: {output_xlsx}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 auto_excel.py <image_path>")
        sys.exit(1)
    process_image(sys.argv[1])
