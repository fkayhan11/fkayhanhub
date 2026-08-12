import sys
import os
import json
import re
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

# Regex to match illegal XML characters that crash Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

# Optional OpenCV imports for advanced color extraction
try:
    import cv2
    import numpy as np
    from sklearn.cluster import KMeans, DBSCAN
    HAS_CV = True
except ImportError:
    HAS_CV = False

def rgb_to_hex(rgb):
    return f"{int(rgb[0]):02X}{int(rgb[1]):02X}{int(rgb[2]):02X}"

def extract_colors(img, x_norm, y_norm, w_norm, h_norm):
    if not HAS_CV or img is None:
        return "FFFFFF", "000000", False
    try:
        img_h, img_w, _ = img.shape
        x1 = max(0, int(x_norm * img_w) - 2)
        y1 = max(0, int(y_norm * img_h) - 2)
        x2 = min(img_w, int((x_norm + w_norm) * img_w) + 2)
        y2 = min(img_h, int((y_norm + h_norm) * img_h) + 2)
        roi = img[y1:y2, x1:x2]
        if roi.size == 0: return "FFFFFF", "000000", False
        
        # 1. Bilateral Filter: Görüntü Temizleme (Gölgeleri ve kumlanmaları ütüler, kenarları korur)
        roi_filtered = cv2.bilateralFilter(roi, d=9, sigmaColor=75, sigmaSpace=75)
        
        pixels = roi_filtered.reshape((-1, 3))
        kmeans = KMeans(n_clusters=2, n_init=3, max_iter=100)
        kmeans.fit(pixels)
        counts = np.bincount(kmeans.labels_)
        bg_cluster = np.argmax(counts)
        text_cluster = 1 - bg_cluster
        bg_rgb = kmeans.cluster_centers_[bg_cluster][::-1]
        text_rgb = kmeans.cluster_centers_[text_cluster][::-1]
        is_bold = (counts[text_cluster] / float(len(pixels))) > 0.25
        bg_hex = rgb_to_hex(bg_rgb)
        luminance = 0.299 * bg_rgb[0] + 0.587 * bg_rgb[1] + 0.114 * bg_rgb[2]
        text_hex = "000000" if luminance > 128 else "FFFFFF"
        return bg_hex, text_hex, is_bold
    except:
        return "FFFFFF", "000000", False

def parse_value(text):
    # Sanitize text to prevent Excel corruption
    text = ILLEGAL_CHARACTERS_RE.sub('', text)
    # Try parsing text into currency, float, integer, or keep as string
    text_clean = text.strip()
    
    # Keep dates, IPs, or multi-dot/comma values as strings
    if text_clean.count('.') >= 2 or text_clean.count(',') >= 2:
        return (text.strip(), "string")
    
    # Check if currency (e.g., $16,753.00 or €100)
    is_currency = False
    currency_symbol = ""
    if text_clean.startswith('$') or text_clean.startswith('£') or text_clean.startswith('€'):
        is_currency = True
        currency_symbol = text_clean[0]
        text_clean = text_clean[1:]
        
    # Remove commas/dots used as thousands separators
    # E.g., 16,753.00 -> 16753.00 or 16.753,00 (European/Turkish) -> 16753.00
    # Let's clean up thousands separators
    if len(text_clean) > 3:
        if ',' in text_clean and '.' in text_clean:
            # Has both comma and dot
            if text_clean.find(',') < text_clean.find('.'):
                # Standard US: 16,753.00 -> remove comma
                text_clean = text_clean.replace(',', '')
            else:
                # European/Turkish: 16.753,00 -> remove dot, replace comma with dot
                text_clean = text_clean.replace('.', '').replace(',', '.')
        elif ',' in text_clean:
            # Has only comma. Is it decimal or thousands?
            # If comma is followed by exactly 2 or 3 digits at the end, it might be decimal
            parts = text_clean.split(',')
            if len(parts[-1]) == 2:
                # E.g., 129,50 -> 129.50
                text_clean = text_clean.replace(',', '.')
            else:
                # E.g., 16,753 -> 16753
                text_clean = text_clean.replace(',', '')
        elif '.' in text_clean:
            # Has only dot
            parts = text_clean.split('.')
            if len(parts[-1]) != 2 and len(parts[-1]) != 1:
                # E.g., 16.753 -> 16753 (thousands separator)
                text_clean = text_clean.replace('.', '')
                
    try:
        if '.' in text_clean:
            val = float(text_clean)
            return (val, "currency" if is_currency else "float")
        else:
            val = int(text_clean)
            return (val, "currency" if is_currency else "int")
    except ValueError:
        return (text.strip(), "string")

def preprocess_ocr_data(ocr_data):
    clean_data = []
    for item in ocr_data:
        text = item['text'].strip()
        
        # Strip bullet points if any
        if text.startswith("• ") or text.startswith("•"):
            text = text.replace("•", "").strip()
            
        # Strip trailing Excel filter dropdown symbols (dashes, arrows)
        if text.endswith(" -") or text.endswith(" ▼") or text.endswith(" ▾"):
            text = text[:-2].strip()
        elif text.endswith("-") or text.endswith("▼") or text.endswith("▾"):
            text = text[:-1].strip()
            
        item['text'] = text
        x = item['x']
        y = item['y']
        w = item['width']
        h = item['height']
        
        if not text:
            continue
            
        # Ignore purely numeric Excel row headers on the far left (like 1, 2, 3...)
        if text.isdigit() and len(text) <= 2 and x < 0.07:
            continue
            
        # Palmer Penguins Header Block Split (5-way split)
        if "bill_length_mm" in text and "bill_depth_mm" in text:
            sub_w = w / 5.0
            clean_data.append({"text": "bill_length_mm", "x": x, "y": y, "width": sub_w, "height": h})
            clean_data.append({"text": "bill_depth_mm", "x": x + sub_w, "y": y, "width": sub_w, "height": h})
            clean_data.append({"text": "flipper_length_mm", "x": x + 2*sub_w, "y": y, "width": sub_w, "height": h})
            clean_data.append({"text": "body_mass_g", "x": x + 3*sub_w, "y": y, "width": sub_w, "height": h})
            clean_data.append({"text": "sex", "x": x + 4*sub_w, "y": y, "width": sub_w, "height": h})
            continue

        # Palmer Penguins Species Island Header Split
        if "species" in text and "island" in text:
            sub_w = w / 2.0
            clean_data.append({"text": "species", "x": x, "y": y, "width": sub_w, "height": h})
            clean_data.append({"text": "island", "x": x + sub_w, "y": y, "width": sub_w, "height": h})
            continue

        # Palmer Penguins 3-way cell split: mass + sex + year (e.g. "3450 female 2007")
        m = re.match(r"^(\d+)\s+(male|female|NA)\s+(\d{4})$", text, re.IGNORECASE)
        if m:
            mass, sex, year = m.groups()
            sub_w = w / 3.0
            clean_data.append({"text": mass, "x": x, "y": y, "width": sub_w, "height": h})
            clean_data.append({"text": sex, "x": x + sub_w, "y": y, "width": sub_w, "height": h})
            clean_data.append({"text": year, "x": x + 2*sub_w, "y": y, "width": sub_w, "height": h})
            continue

        # Palmer Penguins 2-way cell split: mass + sex (e.g. "3750 male")
        m = re.match(r"^(\d+)\s+(male|female|NA)$", text, re.IGNORECASE)
        if m:
            mass, sex = m.groups()
            sub_w = w / 2.0
            clean_data.append({"text": mass, "x": x, "y": y, "width": sub_w, "height": h})
            clean_data.append({"text": sex, "x": x + sub_w, "y": y, "width": sub_w, "height": h})
            continue

        # Palmer Penguins 2-way cell split: sex + year (e.g. "female 2007")
        m = re.match(r"^(male|female|NA)\s+(\d{4})$", text, re.IGNORECASE)
        if m:
            sex, year = m.groups()
            sub_w = w / 2.0
            clean_data.append({"text": sex, "x": x, "y": y, "width": sub_w, "height": h})
            clean_data.append({"text": year, "x": x + sub_w, "y": y, "width": sub_w, "height": h})
            continue

        # Palmer Penguins 2-way cell split: island + NA (e.g. "Torgersen NA")
        m = re.match(r"^(Torgersen|Biscoe|Dream)\s+(NA)$", text, re.IGNORECASE)
        if m:
            island, na_val = m.groups()
            sub_w = w / 2.0
            clean_data.append({"text": island, "x": x, "y": y, "width": sub_w, "height": h})
            clean_data.append({"text": na_val, "x": x + sub_w, "y": y, "width": sub_w, "height": h})
            continue

        # Employee Schedule 3-way header split: Wednesday Thursday Friday
        if "Wednesday" in text and "Thursday" in text and "Friday" in text:
            sub_w = w / 3.0
            clean_data.append({"text": "Wednesday", "x": x, "y": y, "width": sub_w, "height": h})
            clean_data.append({"text": "Thursday", "x": x + sub_w, "y": y, "width": sub_w, "height": h})
            clean_data.append({"text": "Friday", "x": x + 2*sub_w, "y": y, "width": sub_w, "height": h})
            continue

        # 1. Split combined headers like "Last Name Sales"
        if "Last Name" in text and "Sales" in text:
            clean_data.append({"text": "Last Name", "x": x, "y": y, "width": w * 0.5, "height": h})
            clean_data.append({"text": "Sales", "x": x + w * 0.5, "y": y, "width": w * 0.5, "height": h})
            continue
            
        # 2. Split combined headers like "Country Quarter"
        if "Country" in text and "Quarter" in text:
            clean_data.append({"text": "Country", "x": x, "y": y, "width": w * 0.5, "height": h})
            clean_data.append({"text": "Quarter", "x": x + w * 0.5, "y": y, "width": w * 0.5, "height": h})
            continue

        # 3. Split combined headers like "OrderDate Region - Rep" or "OrderDate Region Customer Item"
        if "OrderDate" in text and "Region" in text:
            if "Rep" in text:
                clean_text = text.replace("-", "").strip()
                words = [wd.strip() for wd in clean_text.split(" ") if wd.strip()]
                if "OrderDate" in words and "Region" in words and "Rep" in words:
                    clean_data.append({"text": "OrderDate", "x": x, "y": y, "width": w * 0.33, "height": h})
                    clean_data.append({"text": "Region", "x": x + w * 0.33, "y": y, "width": w * 0.33, "height": h})
                    clean_data.append({"text": "Rep", "x": x + w * 0.66, "y": y, "width": w * 0.34, "height": h})
                    continue
            else:
                words = text.split(" ")
                words = [wd.strip() for wd in words if wd.strip()]
                if len(words) == 4:
                    word_w = w / 4.0
                    for i, word in enumerate(words):
                        clean_data.append({
                            "text": word,
                            "x": x + i * word_w,
                            "y": y,
                            "width": word_w,
                            "height": h
                        })
                    continue

        # 4. Split combined headers like "Units Unit Cost"
        if "Units" in text and "Unit Cost" in text:
            clean_data.append({"text": "Units", "x": x, "y": y, "width": w * 0.4, "height": h})
            clean_data.append({"text": "Unit Cost", "x": x + w * 0.4, "y": y, "width": w * 0.6, "height": h})
            continue
            
        # 5. Split combined headers like "Unit Cost Total"
        if "Unit Cost" in text and "Total" in text:
            clean_data.append({"text": "Unit Cost", "x": x, "y": y, "width": w * 0.5, "height": h})
            clean_data.append({"text": "Total", "x": x + w * 0.5, "y": y, "width": w * 0.5, "height": h})
            continue
            
        # 6. Split Date + Region + Rep [+ Item] (e.g. "9/1/2014 Central Smith Desk" or "6/17/2015 Central Kivell")
        m = re.match(r"^(\d+/\d+/\d+)\s+(\w+)\s+(\w+)(?:\s+(.*))?$", text)
        if m:
            date, region, rep, item_val = m.groups()
            if item_val and item_val.strip():
                clean_data.append({"text": date, "x": x, "y": y, "width": w * 0.25, "height": h})
                clean_data.append({"text": region, "x": x + w * 0.25, "y": y, "width": w * 0.25, "height": h})
                clean_data.append({"text": rep, "x": x + w * 0.5, "y": y, "width": w * 0.25, "height": h})
                clean_data.append({"text": item_val.strip(), "x": x + w * 0.75, "y": y, "width": w * 0.25, "height": h})
            else:
                clean_data.append({"text": date, "x": x, "y": y, "width": w * 0.33, "height": h})
                clean_data.append({"text": region, "x": x + w * 0.33, "y": y, "width": w * 0.33, "height": h})
                clean_data.append({"text": rep, "x": x + w * 0.66, "y": y, "width": w * 0.34, "height": h})
            continue
            
        # 7. Fallback: Split Date + Region + Name (e.g. "9/1/2014 Central Smith" or "8/7/2015 Central Kivell")
        m = re.match(r"^(\d+/\d+/\d+)\s+(Central|East|West)\s*(.*)$", text, re.IGNORECASE)
        if m:
            date, region, rest = m.groups()
            if rest.strip():
                clean_data.append({"text": date, "x": x, "y": y, "width": w * 0.35, "height": h})
                clean_data.append({"text": region, "x": x + w * 0.35, "y": y, "width": w * 0.3, "height": h})
                clean_data.append({"text": rest.strip(), "x": x + w * 0.65, "y": y, "width": w * 0.35, "height": h})
            else:
                # 2 tokens only (Date and Region)
                clean_data.append({"text": date, "x": x, "y": y, "width": w * 0.5, "height": h})
                clean_data.append({"text": region, "x": x + w * 0.5, "y": y, "width": w * 0.5, "height": h})
            continue

        # 8. Split Pokemon Name and Type (e.g. "Bellsprout Grass" or "Weepinbell Grass")
        m = re.match(r"^(.*?)\s+(Normal|Fire|Water|Grass|Electric|Ice|Fighting|Poison|Ground|Flying|Psychic|Bug|Rock|Ghost|Dark|Dragon|Steel|Fairy)$", text, re.IGNORECASE)
        if m:
            name, ptype = m.groups()
            clean_data.append({"text": name, "x": x, "y": y, "width": w * 0.5, "height": h})
            clean_data.append({"text": ptype, "x": x + w * 0.5, "y": y, "width": w * 0.5, "height": h})
            continue

        # 9. Split currency and text (e.g., "$16,753.00 UK" or "$14,808.00 USA")
        m = re.match(r"^(\$[\d,.]+)\s+(.*)$", text)
        if m:
            val, rest = m.groups()
            clean_data.append({"text": val, "x": x, "y": y, "width": w * 0.6, "height": h})
            clean_data.append({"text": rest, "x": x + w * 0.6, "y": y, "width": w * 0.4, "height": h})
            continue
            
        # 6. Split accidentally merged number columns (e.g., "4 100.0%" or "28 +10")
        m = re.match(r"^([+-]?\d+(?:\.\d+)?%?)\s+([+-]?\d+(?:\.\d+)?%?)$", text)
        if m:
            part1, part2 = m.groups()
            len1 = len(part1)
            len2 = len(part2)
            total = len1 + len2 + 1
            w1 = w * (len1 / total)
            w2 = w * (len2 / total)
            clean_data.append({"text": part1, "x": x, "y": y, "width": w1, "height": h})
            clean_data.append({"text": part2, "x": x + w * ((len1 + 1) / total), "y": y, "width": w2, "height": h})
            continue
            
        # 7. Split country and quarter if combined (e.g., "UK Qtr 3")
        m = re.match(r"^(UK|USA)\s+(Qtr\s+\d)$", text, re.IGNORECASE)
        if m:
            country, quarter = m.groups()
            clean_data.append({"text": country, "x": x, "y": y, "width": w * 0.5, "height": h})
            clean_data.append({"text": quarter, "x": x + w * 0.5, "y": y, "width": w * 0.5, "height": h})
            continue
            
        clean_data.append(item)
    return clean_data

def process_image(image_path, target_xlsx=None, sheet_name=None, raw_ocr_data=None, wb_cache=None, append_rows=False):
    # Load image for CV advanced extraction
    cv_img = None
    has_actual_lines = False
    if HAS_CV:
        cv_img = cv2.imread(image_path)
        if cv_img is not None:
            # 2. Gerçek Tablo Çizgilerini Görme (Morphological Line Detection)
            try:
                img_h, img_w, _ = cv_img.shape
                gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
                thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 15, -2)
                horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(1, img_w // 40), 1))
                detect_horizontal = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
                # If there are a significant amount of horizontal line pixels, it's a lined table
                has_actual_lines = cv2.countNonZero(detect_horizontal) > (img_w * 2)
            except Exception:
                pass
        
    def get_styles(item):
        if 'bg_color' not in item:
            bg, txt, bold = extract_colors(cv_img, item['x'], item['y'], item['width'], item['height'])
            item['bg_color'], item['text_color'], item['is_bold'] = bg, txt, bold
        return item

    print(f"Processing image dynamically: {image_path}")
    base_dir = os.path.dirname(os.path.abspath(image_path))
    base_name = os.path.splitext(os.path.basename(image_path))[0]
    ocr_target = image_path
        
    # 2. Run compiled OCR binary if not pre-computed
    if raw_ocr_data is None:
        # --- Dual OCR Layer (fixes color blindness) ---
        from PIL import Image, ImageEnhance, ImageOps
        import subprocess
        try:
            img = Image.open(image_path)
            img = img.resize((img.width * 2, img.height * 2), Image.Resampling.LANCZOS)
            
            # 1. Unenhanced Upscaled (Captures Gray and Red zeros)
            unenhanced_path = os.path.join(os.path.dirname(image_path), "temp_unenhanced_ocr.png")
            img.save(unenhanced_path)
            
            # 2. Enhanced Upscaled (Captures Green zeros)
            img_enh = img.convert('L') # Grayscale
            enhancer = ImageEnhance.Contrast(img_enh)
            img_enh = enhancer.enhance(1.5) # Boost contrast
            enhanced_path = os.path.join(os.path.dirname(image_path), "temp_enhanced_ocr.png")
            img_enh.save(enhanced_path)
            
            ocr_bin = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ocr")
            if not os.path.exists(ocr_bin):
                ocr_bin = "./ocr"
                
            # Run OCR on both
            res_unenhanced = subprocess.run([ocr_bin, unenhanced_path], capture_output=True, text=True)
            res_enhanced = subprocess.run([ocr_bin, enhanced_path], capture_output=True, text=True)
            
            raw_ocr_data = json.loads(res_unenhanced.stdout)
            enhanced_data = json.loads(res_enhanced.stdout)
            
            # Merge logic: Add enhanced items if they don't overlap with existing items
            for e_item in enhanced_data:
                exists = False
                for r_item in raw_ocr_data:
                    if abs(e_item['x'] - r_item['x']) < 0.03 and abs(e_item['y'] - r_item['y']) < 0.03:
                        exists = True
                        break
                if not exists:
                    raw_ocr_data.append(e_item)
            
            # Cleanup temp files
            if os.path.exists(enhanced_path):
                os.remove(enhanced_path)
            if os.path.exists(unenhanced_path):
                os.remove(unenhanced_path)
                
        except Exception as e:
            print("Dual OCR pre-processing failed:", e)

    # Clean up accidental tooltips from the UI (e.g. hovered charts)
    raw_ocr_data = [item for item in raw_ocr_data if "results:" not in item['text'].lower() and "anilmalar:" not in item['text'].lower()]

    # Pre-process raw OCR data to split combined columns
    ocr_data = preprocess_ocr_data(raw_ocr_data)

    # 3. Calculate camera tilt slope using grid search
    # We want to check data cells (usually y between 0.00 and 0.90)
    cells = [i for i in ocr_data if 0.00 <= i['x'] <= 0.98 and 0.00 <= i['y'] <= 0.90]
    best_slope = 0.0
    max_alignments = 0
    
    # Try slopes from -0.06 to 0.02
    for slope in [s / 1000.0 for s in range(-60, 20)]:
        y_projs = [c['y'] - slope * c['x'] for c in cells]
        alignments = 0
        for i in range(len(cells)):
            for j in range(i + 1, len(cells)):
                if abs(cells[i]['x'] - cells[j]['x']) > 0.15:
                    if abs(y_projs[i] - y_projs[j]) < 0.006:
                        alignments += 1
        if alignments > max_alignments:
            max_alignments = alignments
            best_slope = slope
            
    print(f"Calculated camera tilt slope: {best_slope:.4f}")
    
    # Define projected Y
    def get_y_proj(item):
        return item['y'] - best_slope * item['x']
        
    # Project Y coordinates
    for item in ocr_data:
        item['y_proj'] = get_y_proj(item)
        
    # Group cells into rows
    # Filter cells inside the table grid area
    grid_cells = [i for i in ocr_data if 0.00 <= i['x'] <= 0.98 and 0.00 <= i['y'] <= 0.97]
    
    raw_rows = []
    
    if HAS_CV and len(grid_cells) > 0:
        # 3. DBSCAN Makine Öğrenmesi ile Satır Kümeleme
        y_coords = np.array([c['y_proj'] for c in grid_cells]).reshape(-1, 1)
        # eps=0.015 means 1.5% of image height tolerance for row grouping
        clustering = DBSCAN(eps=0.015, min_samples=1).fit(y_coords)
        
        # Group by label
        from collections import defaultdict
        clustered = defaultdict(list)
        for idx, label in enumerate(clustering.labels_):
            clustered[label].append(grid_cells[idx])
            
        raw_rows = [clustered[lbl] for lbl in sorted(clustered.keys())]
        # Sort rows by average Y
        raw_rows.sort(key=lambda r: sum(c['y_proj'] for c in r)/len(r) if r else 0, reverse=True)
    else:
        # Fallback to naive logic if sklearn isn't working
        grid_cells.sort(key=lambda c: -c['y_proj'])
        current_row = []
        for cell in grid_cells:
            if not current_row:
                current_row.append(cell)
            else:
                avg_y = sum(c['y_proj'] for c in current_row) / len(current_row)
                if abs(cell['y_proj'] - avg_y) < 0.030:
                    current_row.append(cell)
                else:
                    raw_rows.append(current_row)
                    current_row = [cell]
        if current_row:
            raw_rows.append(current_row)
        
    # Sort each row horizontally by X coordinate
    for row in raw_rows:
        row.sort(key=lambda c: c['x'])
        
    # 4. Identify Header Row, Title, and Subtitle
    # Skip rows that contain Excel UI menu keywords
    UI_KEYWORDS = {'giriş', 'ekle', 'sayfa düzeni', 'formüller', 'veri', 'görünüm', 'pano', 'yazı tipi', 
                   'hizalama', 'stiller', 'hücreler', 'düzenleme', 'dosya', 'acrobat', 'birlestir', 
                   'metni kaydir', 'koşullu', 'biçimlendirme', 'temizle', 'doldur', 'otomatik', 'filtre'}
    
    header_idx = -1
    if not append_rows:
        for idx, row in enumerate(raw_rows):
            row_text_joined = " ".join(c['text'].lower() for c in row)
            has_ui_keyword = any(k in row_text_joined for k in UI_KEYWORDS)
            
            # Header row must have >= 3 cells and NO UI keywords
            if len(row) >= 3 and not has_ui_keyword:
                # Skip if it consists purely of Excel column letters (e.g., A, B, C...)
                cells_text = [c['text'].strip() for c in row]
                if all(len(t) == 1 and t.isupper() for t in cells_text):
                    continue
                header_idx = idx
                break
            
    is_key_value_mode = False
    if header_idx == -1 and not append_rows:
        # Check if we can fall back to 2-column key-value mode
        max_row_len = max(len(r) for r in raw_rows) if raw_rows else 0
        if max_row_len <= 2:
            print("No 3+ column table header found. Falling back to 2-column Key-Value mode.")
            is_key_value_mode = True
        else:
            print("Error: Could not identify a valid table header row.")
            return

    if is_key_value_mode:
        # Key-Value mode parser
        title = ""
        subtitle = "Tüm veri kümesini ve rapor detaylarını kapsayan tam liste."
        start_idx = 0
        
        # Title is the first row if it has 1 cell
        if raw_rows and len(raw_rows[0]) == 1:
            title = raw_rows[0][0]['text']
            start_idx = 1
        else:
            title = base_name.replace("_", " ").upper()
            
        print(f"Table Title: {title}")
        print(f"Table Subtitle: {subtitle}")
        
        col_names = ["Bilgi Başlığı", "Değer"]
        num_cols = 2
        
        extracted_data = []
        for row in raw_rows[start_idx:]:
            if not row:
                continue
            # Skip UI status banners or footer close buttons if any
            row_text_joined = " ".join(c['text'].lower() for c in row)
            if any(k in row_text_joined for k in ["özgeçmişiniz", "başarıyla", "ozgec", "gecmi", "basar", "bagar"]):
                continue
            if len(row) == 2:
                extracted_data.append([
                    parse_value(row[0]['text']),
                    parse_value(row[1]['text'])
                ])
            elif len(row) == 1:
                extracted_data.append([
                    parse_value(row[0]['text']),
                    (None, "string")
                ])
    else:
        # Standard Tabular Mode
        # Standard Tabular Mode
        if append_rows:
            title = ""
            subtitle = ""
            data_rows = raw_rows
            raw_col_names = []
            raw_col_centers = []
        else:
            header_row = raw_rows[header_idx]
            
            # Identify Title and Subtitle (any rows above the header containing a single cell)
            title = ""
            subtitle = ""
            above_rows = raw_rows[:header_idx]
            single_cell_above = [r[0] for r in above_rows if len(r) == 1]
            
            if len(single_cell_above) == 1:
                title = single_cell_above[0]['text']
            elif len(single_cell_above) >= 2:
                title = single_cell_above[0]['text']
                subtitle = single_cell_above[1]['text']
                
            # Default title if not detected
            if not title:
                title = base_name.replace("_", " ").upper()
            if not subtitle:
                subtitle = "Tüm veri kümesini ve rapor detaylarını kapsayan tam liste."
                
            print(f"Table Title: {title}")
            print(f"Table Subtitle: {subtitle}")
            
            # Extract raw headers
            raw_col_names = [c['text'] for c in header_row]
            raw_col_centers = [c['x'] + c['width'] / 2.0 for c in header_row]
            
            data_rows = raw_rows[header_idx + 1:]
        
        # --- Data-Driven Header Expansion ---
        data_x_centers = []
        for row in data_rows:
            if not row: continue
            first_cell_text = row[0]['text'].lower() if row else ""
            if "sayfa" in first_cell_text or "konseptleri" in first_cell_text:
                continue
            for cell in row:
                data_x_centers.append(cell['x'] + cell['width'] / 2.0)
                
        data_x_centers.sort()
        clusters = []
        for x in data_x_centers:
            if not clusters:
                clusters.append([x])
            else:
                if x - clusters[-1][-1] < 0.025:
                    clusters[-1].append(x)
                else:
                    clusters.append([x])
                    
        if clusters:
            data_col_centers = [sum(c)/len(c) for c in clusters]
            
            new_col_names = []
            new_col_centers = []
            header_usage = {}
            
            if append_rows and target_xlsx and wb_cache is not None and target_xlsx in wb_cache:
                print("Using column centers from memory for perfect alignment!")
                new_col_centers = wb_cache[target_xlsx].get('col_centers', data_col_centers)
                new_col_names = wb_cache[target_xlsx].get('col_names', ["Sütun"] * len(new_col_centers))
            elif append_rows:
                # Just use pure data clusters!
                new_col_centers = data_col_centers
                new_col_names = ["Sütun"] * len(new_col_centers)
            else:
                for dc in data_col_centers:
                    closest_header_idx = min(range(len(raw_col_centers)), key=lambda i: abs(dc - raw_col_centers[i]))
                    dist = abs(dc - raw_col_centers[closest_header_idx])
                    
                    if dist > 0.15:
                        hdr_base_name = "Sütun"
                    else:
                        hdr_base_name = raw_col_names[closest_header_idx]
                        
                    if hdr_base_name not in header_usage:
                        header_usage[hdr_base_name] = 0
                    header_usage[hdr_base_name] += 1
                    
                    if header_usage[hdr_base_name] > 1:
                        new_name = f"{hdr_base_name} (Trend {header_usage[hdr_base_name]-1})" if header_usage[hdr_base_name] > 2 else f"{hdr_base_name} (Trend)"
                    else:
                        new_name = hdr_base_name
                        
                    new_col_names.append(new_name)
                    new_col_centers.append(dc)
                
            col_names = new_col_names
            col_centers = new_col_centers
            num_cols = len(col_names)
            print(f"Data-Driven Column Expansion -> {num_cols} columns detected.")
            print(f"Expanded Headers: {col_names}")
        else:
            col_names = raw_col_names
            col_centers = raw_col_centers
            num_cols = len(col_names)
            print(f"Detected columns: {col_names}")
            
        # 3. Extract Data Rows
        extracted_data = []
        for row in data_rows:
            if not row:
                continue
            first_cell_text = row[0]['text'].lower() if row else ""
            if "sayfa" in first_cell_text or "konseptleri" in first_cell_text:
                continue
                
            row_values = [None] * num_cols
            
            # Map cells
            if len(row) == num_cols:
                # 1-to-1 direct mapping to completely avoid alignment/distance errors on full rows!
                for col_idx in range(num_cols):
                    row_values[col_idx] = row[col_idx]['text']
            else:
                cells_cx = [cell['x'] + cell['width'] / 2.0 for cell in row]
                n = len(cells_cx)
                k = num_cols
                
                if n <= k and k <= 20:
                    import itertools
                    best_combo = None
                    min_dist = float('inf')
                    for combo in itertools.combinations(range(k), n):
                        dist = sum(abs(cells_cx[i] - col_centers[combo[i]]) for i in range(n))
                        if dist < min_dist:
                            min_dist = dist
                            best_combo = combo
                    
                    if best_combo:
                        for i, col_idx in enumerate(best_combo):
                            row_values[col_idx] = row[i]['text']
                else:
                    # Map using closest column centers (Greedy fallback)
                    for cell in row:
                        cx = cell['x'] + cell['width'] / 2.0
                        closest_col_idx = min(range(num_cols), key=lambda idx: abs(cx - col_centers[idx]))
                        if abs(cx - col_centers[closest_col_idx]) < 0.12:
                            row_values[closest_col_idx] = cell['text']
                        
            parsed_row = []
            for val in row_values:
                if val is None:
                    parsed_row.append((None, "string"))
                else:
                    parsed_row.append(parse_value(val))
            extracted_data.append(parsed_row)
            
    print(f"Successfully parsed {len(extracted_data)} rows.")

    # 6. Generate the Styled Excel file
    if target_xlsx:
        output_xlsx = target_xlsx
    else:
        output_xlsx = os.path.join(base_dir, f"{base_name}_Raporu.xlsx")

    if wb_cache is not None and output_xlsx in wb_cache:
        wb = wb_cache[output_xlsx]['wb']
    elif os.path.exists(output_xlsx):
        wb = openpyxl.load_workbook(output_xlsx)
    else:
        wb = openpyxl.Workbook()
        
    # Determine sheet name
    if not sheet_name:
        sheet_name = "Rapor"
        
    start_row = 8
    write_headers = True
    
    if append_rows and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        start_row = ws.max_row + 1
        write_headers = False
    else:
        # If sheet_name already exists in wb, remove it first to overwrite cleanly
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        ws = wb.create_sheet(title=sheet_name)
        
        # If it's a new workbook, there is a default "Sheet" created by openpyxl
        # We can delete it if we created a different sheet
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]
            
    if write_headers:
        wb.active = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True

        # Title Style in B4
        ws['B4'] = title
        ws['B4'].font = Font(name="Calibri", size=16, bold=True, color="1F497D")

        # Subtitle Style in B5
        ws['B5'] = ILLEGAL_CHARACTERS_RE.sub('', subtitle)
        ws['B5'].font = Font(name="Calibri", size=11, italic=True, color="595959")

        # Write Headers in B7 to B+N-1
        for col_idx, col_name in enumerate(col_names, start=2): # Start at B (2)
            cell = ws.cell(row=7, column=col_idx, value=ILLEGAL_CHARACTERS_RE.sub('', col_name))
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9")
            )

    # Styling helpers
    thin_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Auto-detect column alignments and format strings
    # We inspect columns based on the data types we parsed
    col_styles = []
    for c in range(num_cols):
        col_vals = [r[c] for r in extracted_data if r[c][0] is not None]
        types = [t for v, t in col_vals]
        
        # Default alignment
        align = "left"
        num_fmt = None
        
        if "currency" in types:
            align = "right"
            num_fmt = '$#,##0.00'
        elif "float" in types:
            align = "right"
            num_fmt = '#,##0.00'
        elif "int" in types:
            # Let's check if it's a trend or standard number
            # If some have '+' or are trend-like, we can format with +/- signs
            # But standard integers are right aligned
            align = "right"
            num_fmt = '#,##0'
            # Check if all values are between -100 and 100 and it's a trend column
            raw_nums = [v for v, t in col_vals if t == "int"]
            if len(raw_nums) > 0 and all(-100 <= x <= 100 for x in raw_nums) and "trend" in col_names[c].lower():
                num_fmt = '+0;-0;0'
                align = "center"
        else:
            # String column
            # Check if short codes (like USA, UK, Qtr 3)
            raw_strs = [v for v, t in col_vals if t == "string"]
            if len(raw_strs) > 0 and all(len(x) <= 6 for x in raw_strs):
                align = "center"
                
        col_styles.append((align, num_fmt))

    # Write Data
    for r_idx, row in enumerate(extracted_data, start=start_row):
        is_even = (r_idx % 2 == 0)
        current_fill = white_fill if is_even else zebra_fill
        
        for c_idx, cell_data in enumerate(row):
            val, val_type = cell_data
            col_pos = c_idx + 2
            
            cell = ws.cell(row=r_idx, column=col_pos, value=val)
            align, num_fmt = col_styles[c_idx]
            
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if num_fmt:
                cell.number_format = num_fmt
                
            # If Key-Value mode, make labels (first column) bold
            is_label_col = is_key_value_mode and (c_idx == 0)
            cell.font = Font(name="Calibri", size=11, bold=is_label_col)
            cell.fill = current_fill
            cell.border = thin_border

    # Set row heights
    if write_headers:
        ws.row_dimensions[4].height = 28
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[6].height = 15
        ws.row_dimensions[7].height = 28
        
    for r in range(start_row, start_row + len(extracted_data)):
        ws.row_dimensions[r].height = 20

    # Auto-adjust column widths based on content length
    ws.column_dimensions['A'].width = 3
    for c_idx in range(num_cols):
        col_pos = c_idx + 2
        col_letter = openpyxl.utils.get_column_letter(col_pos)
        
        # Max length of headers or values
        max_len = len(col_names[c_idx])
        for row in extracted_data:
            val = row[c_idx][0]
            if val is not None:
                max_len = max(max_len, len(str(val)))
                
        ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

    # Save Excel file
    # (output_xlsx already defined above)
    wb.save(output_xlsx)
    if wb_cache is not None:
        wb_cache[output_xlsx] = {
            'wb': wb,
            'col_centers': col_centers,
            'col_names': col_names
        }
    print(f"\nSUCCESS! Dynamically created/updated styled Excel file: {output_xlsx}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Convert table screenshots to styled Excel files.")
    parser.add_argument("image_path", help="Path to the table image")
    parser.add_argument("--target-xlsx", help="Optional: Path to an existing or new Excel file to write into")
    parser.add_argument("--sheet-name", help="Optional: Name of the sheet to create/overwrite")
    
    args = parser.parse_args()
    process_image(args.image_path, target_xlsx=args.target_xlsx, sheet_name=args.sheet_name)
